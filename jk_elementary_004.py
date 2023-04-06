from pathlib import Path
from openpyxl import load_workbook
import pandas as pd
import datetime

def get_invoice_data(file_path: Path) -> (datetime.date, str, int):
        wb = load_workbook(file_path, data_only=True)
        ws = wb.worksheets[0]
        inv_datetime = ws["A2"].value
        # 日付
        inv_date =  datetime.date(
            inv_datetime.year,
            inv_datetime.month,
            inv_datetime.day
            ) 
        # 請求先
        inv_company = ws["B2"].value
        # 金額
        inv_amount = ws["C2"].value
        return inv_date, inv_company, inv_amount

def main():
    folder = Path(r"C:\Users\yosak\Desktop\JKSapu\2023")

    df = pd.DataFrame()
    for file_path in folder.glob("*.xlsx"):
        # 請求書データ抽出
        inv_date, inv_com, inv_amo = get_invoice_data(file_path)
        df = df.append({
            "日付": inv_date,
            "請求先": inv_com,
            "金額": inv_amo        
        }, ignore_index=True)
    df.to_excel("2023.xlsx")

main()
